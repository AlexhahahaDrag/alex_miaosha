#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
生成联系人导入模版
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import sys
import io

# 设置输出编码
if sys.platform == 'win32':
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

def generate_template():
    """生成联系人模版 Excel 文件"""
    # 创建新工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = "联系人"  # 将工作表名改为中文

    # 设置中文表头
    headers = ["姓名", "电话", "关系", "邮箱", "地址", "备注"]
    ws.append(headers)

    # 设置表头样式
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=12)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # 边框样式
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # 应用表头样式
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border

    # 添加示例数据（3条）
    sample_data = [
        ["张三", "13812345678", "friend", "zhangsan@example.com", "北京市朝阳区", "高中同学"],
        ["李四", "13912345678", "family", "lisi@example.com", "上海市浦东新区", "表哥"],
        ["王五", "13712345678", "colleague", "wangwu@example.com", "深圳市南山区", "项目经理"],
    ]

    # 应用数据和样式
    data_font = Font(size=11)
    data_alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    for row_data in sample_data:
        ws.append(row_data)
        row_num = ws.max_row
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.font = data_font
            cell.alignment = data_alignment
            cell.border = thin_border

    # 调整列宽
    column_widths = {
        'A': 15,  # 姓名
        'B': 18,  # 电话
        'C': 18,  # 关系
        'D': 25,  # 邮箱
        'E': 25,  # 地址
        'F': 30,  # 备注
    }

    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width

    # 设置行高
    ws.row_dimensions[1].height = 25

    # 保存文件
    file_path = 'contacts_user_template.xlsx'
    wb.save(file_path)
    return file_path

if __name__ == '__main__':
    try:
        file_path = generate_template()
        print("Success: Template file generated successfully")
        print("File: " + file_path)
        print("Content:")
        print("  - Headers: Xingming, Dianhua, Guanxi, Youxiang, Dizhi, Beizhu")
        print("  - Headers (CN): 姓名, 电话, 关系, 邮箱, 地址, 备注")
        print("  - Sample data: 3 records")
        print("  - File encoding: UTF-8")
    except Exception as e:
        print("Error: Failed to generate template file: " + str(e))
        import traceback
        traceback.print_exc()
